"""
Excel Data Importer
Reads water quality data from Excel files
"""

import pandas as pd
from pathlib import Path
from typing import Dict, Optional


class ExcelDataSource:
    """Import water quality data from Excel files"""

    def __init__(self, excel_file: Optional[Path] = None):
        """
        Initialize with optional Excel file path

        Args:
            excel_file: Path to Excel file with water quality data
        """
        self.excel_file = excel_file

    def load_water_quality_data(self, sheet_name: str = "Water_Quality") -> pd.DataFrame:
        """
        Load water quality data from Excel

        Expected Excel structure:
        - Columns: date, borehole_id, industrial_area, parameter, value, unit, lab
        - Rows: Individual test results

        Args:
            sheet_name: Name of sheet containing water quality data

        Returns:
            DataFrame with water quality data
        """
        if not self.excel_file or not self.excel_file.exists():
            print(f"Excel file not found: {self.excel_file}")
            return pd.DataFrame()

        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            df["date"] = pd.to_datetime(df["date"])
            df["value"] = pd.to_numeric(df["value"], errors="coerce")
            return df
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return pd.DataFrame()

    def load_borehole_registry(self, sheet_name: str = "Boreholes") -> pd.DataFrame:
        """
        Load monitoring well registry

        Expected structure:
        - Columns: borehole_id, industrial_area, latitude, longitude, depth_m, established_year

        Args:
            sheet_name: Name of sheet with borehole information

        Returns:
            DataFrame with borehole details
        """
        if not self.excel_file or not self.excel_file.exists():
            return pd.DataFrame()

        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            return df
        except Exception as e:
            print(f"Error reading boreholes sheet: {e}")
            return pd.DataFrame()

    def load_standards(self, sheet_name: str = "Standards") -> Dict[str, float]:
        """
        Load drinking water standards

        Expected structure:
        - Columns: parameter, value, unit

        Args:
            sheet_name: Name of sheet with standards

        Returns:
            Dictionary mapping parameter names to threshold values
        """
        if not self.excel_file or not self.excel_file.exists():
            return {}

        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            return dict(zip(df["parameter"], df["value"]))
        except Exception as e:
            print(f"Error reading standards: {e}")
            return {}

    def get_area_summary(self, area_name: str) -> Dict:
        """Get summary statistics for industrial area from Excel data"""
        df = self.load_water_quality_data()

        if df.empty or "industrial_area" not in df.columns:
            return {}

        area_data = df[df["industrial_area"].str.lower() == area_name.lower()]

        if area_data.empty:
            return {}

        return {
            "area": area_name,
            "total_tests": len(area_data),
            "parameters_tested": area_data["parameter"].unique().tolist(),
            "date_range": {
                "start": area_data["date"].min().strftime("%Y-%m-%d"),
                "end": area_data["date"].max().strftime("%Y-%m-%d")
            },
            "boreholes": area_data["borehole_id"].nunique(),
            "labs_involved": area_data["lab"].unique().tolist() if "lab" in area_data.columns else []
        }

    @staticmethod
    def create_template_excel(output_path: Path) -> None:
        """Create a template Excel file for water quality data input"""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Sheet 1: Water Quality Data
        ws1 = wb.active
        ws1.title = "Water_Quality"
        headers_wq = ["date", "borehole_id", "industrial_area", "parameter", "value", "unit", "lab", "notes"]
        ws1.append(headers_wq)

        sample_rows = [
            ["2023-01-15", "RAIN-01", "raanana", "TCE", 250, "μg/L", "Lab A", "High concentration"],
            ["2023-01-15", "RAIN-01", "raanana", "Chlorides", 450, "mg/L", "Lab A", ""],
            ["2023-02-20", "RAIN-01", "raanana", "TCE", 280, "μg/L", "Lab A", "Increasing trend"],
        ]

        for row in sample_rows:
            ws1.append(row)

        # Sheet 2: Boreholes Registry
        ws2 = wb.create_sheet("Boreholes")
        headers_bh = ["borehole_id", "industrial_area", "latitude", "longitude", "depth_m", "established_year", "status"]
        ws2.append(headers_bh)

        sample_boreholes = [
            ["RAIN-01", "raanana", 32.1950, 34.8639, 45, 2007, "Active"],
            ["RAIN-02", "raanana", 32.1960, 34.8650, 50, 2008, "Active"],
        ]

        for row in sample_boreholes:
            ws2.append(row)

        # Sheet 3: Standards
        ws3 = wb.create_sheet("Standards")
        headers_std = ["parameter", "value", "unit", "hebrew_name"]
        ws3.append(headers_std)

        standards = [
            ["TCE", 8.3, "μg/L", "טריכלורואתילן"],
            ["PCE", 10, "μg/L", "פרכלורואתילן"],
            ["Benzene", 1, "μg/L", "בנזן"],
            ["Chlorides", 250, "mg/L", "כלורידים"],
            ["Nitrates", 50, "mg/L", "ניטרטים"],
        ]

        for row in standards:
            ws3.append(row)

        wb.save(output_path)
        print(f"Template Excel file created: {output_path}")
